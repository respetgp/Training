import time
import logging
from aiogram import Bot, Dispatcher, executor, types
import openpyxl as op
import random

logging.basicConfig(level=logging.INFO)

TOKEN = '5909034210:AAH3cL4uut3i-twYYGBdmVbpUfklQqjl1e8'
# Выбираем файл формата эксель с таблицей
file_name = 'my_artikel.xlsx'
wort_list = []
wb = op.load_workbook(file_name, data_only=True)
sheet = wb.active
# создаем переменную равную числу строк в табцлице
max_rows = sheet.max_row

for i in range(2, max_rows + 1):
    index = sheet.cell(row=i, column=1).value  # приставеваем переменной столбец индекс
    artikel = sheet.cell(row=i, column=2).value  # приставеваем переменной столбец артикль
    wort = sheet.cell(row=i, column=3).value  # присваиваем переменной столбец слово
    translate = sheet.cell(row=i, column=4).value  # приставеваем переменной столбец перевод

    wort_list.append([artikel, wort, translate])

    if not wort:
        continue

#print(*MESSAGE, sep=' ')

bot = Bot(token = TOKEN)
dp = Dispatcher(bot = bot)

@dp.message_handler(commands = 'start')
async def start_handler(message: types.Message):
    user_id = message.from_user.id
    user_name = message.from_user.first_name
    user_full_name = message.from_user.full_name
    logging.info(f'{user_id=} {user_full_name=} {time.asctime()}')
    await message.reply(f'Привет, {user_full_name}!')

    for i in range(15):
        time.sleep(3600)
        MESSAGE = " ".join(random.choice(wort_list))
        await bot.send_message(user_id, MESSAGE.format(user_name))

if __name__ == '__main__':
    executor.start_polling(dp)